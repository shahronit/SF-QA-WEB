"""Load and chunk knowledge-base documents for RAG."""

from __future__ import annotations

import json
import os
from pathlib import Path

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

from langchain_community.document_loaders import (
    CSVLoader,
    JSONLoader,
    PyPDFLoader,
    UnstructuredWordDocumentLoader,
)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


class SalesforceKnowledgeIngestor:
    """Walk `knowledge_base/` and produce chunked LangChain `Document` lists."""

    def __init__(self, knowledge_base_path: str | Path | None = None) -> None:
        root = Path(__file__).resolve().parents[1]
        self.kb_path = Path(knowledge_base_path) if knowledge_base_path else root / "knowledge_base"
        self.splitter = RecursiveCharacterTextSplitter(
            chunk_size=500,
            chunk_overlap=50,
            separators=["\n\n", "\n", ".", " "],
        )

    def load_all(self) -> list[Document]:
        """Load every supported file under the KB path and return split documents."""
        docs: list[Document] = []
        if not self.kb_path.is_dir():
            return self.splitter.split_documents(docs)
        for root, _, files in os.walk(self.kb_path):
            for file in files:
                if file == ".gitkeep":
                    continue
                path = Path(root) / file
                docs.extend(self._load_file(path))
        return self.splitter.split_documents(docs)

    # Extensions whose contents are essentially binary; trying to decode
    # them as UTF-8 produces garbage so the generic text fallback skips
    # them rather than feeding noise into the embedder.
    _BINARY_EXTS: frozenset[str] = frozenset({
        "png", "jpg", "jpeg", "gif", "webp", "bmp", "tiff", "ico", "svgz",
        "zip", "tar", "gz", "bz2", "xz", "7z", "rar",
        "mp3", "mp4", "wav", "ogg", "webm", "avi", "mov", "mkv", "flac",
        "exe", "dll", "so", "dylib", "bin", "iso", "dmg", "pkg",
        "ttf", "otf", "woff", "woff2",
        "ppt", "pptx", "key", "pages", "numbers",
    })

    def _load_file(self, path: Path) -> list[Document]:
        """Load a single file based on its extension.

        Has a generous fallback chain: known structured formats use the
        right loader (PDF/DOCX/CSV/JSON/XLSX), and *every other* readable
        extension is loaded as plain text so the file is still indexed
        for RAG. Truly binary formats (images, archives, executables,
        media) are skipped instead of being decoded into garbage.
        """
        ext = path.suffix.lower().lstrip(".")
        try:
            if ext == "pdf":
                return PyPDFLoader(str(path)).load()
            if ext == "csv":
                return CSVLoader(str(path)).load()
            if ext in ("doc", "docx"):
                return UnstructuredWordDocumentLoader(str(path)).load()
            if ext == "json":
                return self._load_json(path)
            if ext == "xlsx" and load_workbook is not None:
                return self._load_xlsx(path)
            if ext in self._BINARY_EXTS:
                # Skip silently — embedding raw bytes is worse than
                # nothing and the file is still preserved on storage.
                return []
            # Generic text fallback: covers md/txt/markdown PLUS any other
            # text-shaped extension a user might upload (.yaml, .yml,
            # .xml, .html, .htm, .log, .sql, .py, .ts, .js, .ini, .toml,
            # .properties, .env, .rst, .tsv, .conf, …) AND files with no
            # extension at all.
            return self._load_as_text(path)
        except Exception as exc:  # noqa: BLE001 — surface in UI/logs
            print(f"Failed to load {path}: {exc}")
        return []

    def _load_as_text(self, path: Path) -> list[Document]:
        """Read *path* as text and return a single Document.

        Uses ``errors='replace'`` so a stray non-UTF-8 byte doesn't kill
        the entire ingest. Empty files become empty documents which the
        splitter will harmlessly drop.
        """
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except (OSError, UnicodeError) as exc:
            print(f"Failed to read {path} as text: {exc}")
            return []
        if not text.strip():
            return []
        return [Document(page_content=text, metadata={"source": str(path)})]

    def _load_xlsx(self, path: Path) -> list[Document]:
        """Load text from an Excel .xlsx workbook (all sheets, tab-separated rows)."""
        if load_workbook is None:
            return []
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            parts: list[str] = []
            for sheet in wb.worksheets:
                rows: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append("\t".join(cells))
                if rows:
                    parts.append(f"## {sheet.title}\n" + "\n".join(rows))
            wb.close()
            if not parts:
                return [Document(page_content="(empty workbook)", metadata={"source": str(path)})]
            text = "\n\n".join(parts)
            return [Document(page_content=text, metadata={"source": str(path)})]
        except Exception as exc:  # noqa: BLE001
            print(f"Failed to load xlsx {path}: {exc}")
            return []

    def _load_json(self, path: Path) -> list[Document]:
        """Try multiple JSON loading strategies."""
        try:
            return JSONLoader(str(path), jq_schema=".", text_content=False).load()
        except Exception:
            try:
                return JSONLoader(str(path), jq_schema=".[]", text_content=False).load()
            except Exception:
                raw = path.read_text(encoding="utf-8", errors="replace")
                data = json.loads(raw)
                return [Document(page_content=json.dumps(data, indent=2), metadata={"source": str(path)})]
