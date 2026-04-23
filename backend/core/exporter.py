"""Export agent output to Markdown, Excel, CSV, and PDF formats."""

from __future__ import annotations

import csv
import html as _html
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SF_HEADER_FILL = "FF0070D2"
SF_HEADER_HEX = "#0070D2"

_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*#*\s*$")
_FENCE_RE = re.compile(r"^\s*(```|~~~)")

# Match the "<br>" line-break sentinel agents emit inside Markdown table cells,
# whether raw (`<br>`, `<br/>`, `<br />`) or HTML-entity escaped
# (`&lt;br&gt;`, `&lt;br/&gt;`). Used to convert those sentinels to real
# newlines before writing into Excel/CSV cells (where `wrap_text=True` will
# stack them) or to real `<br/>` tags before PDF rendering.
_BR_SENTINEL_RE = re.compile(r"(?:&lt;|<)\s*br\s*/?\s*(?:&gt;|>)", re.IGNORECASE)


def _normalize_cell(value: Any) -> Any:
    """Convert <br>-style sentinels in a cell value to real newlines."""
    if not isinstance(value, str):
        return value
    return _BR_SENTINEL_RE.sub("\n", value)


def export_markdown(content: str, path: str | Path) -> Path:
    """Write raw markdown content to a file and return its path."""
    p = Path(path)
    p.write_text(content, encoding="utf-8")
    return p


def _style_header_row(ws: Any, ncols: int) -> None:
    """Apply Salesforce-blue header styling to the first row."""
    fill = PatternFill(start_color=SF_HEADER_FILL, end_color=SF_HEADER_FILL, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def export_workbook_bytes(sheets: dict[str, list[dict[str, Any]]]) -> bytes:
    """Build an .xlsx with one sheet per key; each value is a list of row dicts (openpyxl)."""
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        name = sheet_name[:31] or "Sheet1"
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        if not rows:
            ws.cell(row=1, column=1, value="(No data)")
            continue
        headers = list(rows[0].keys())
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, len(headers))
        for r, row in enumerate(rows, start=2):
            for c, h in enumerate(headers, start=1):
                val = _normalize_cell(row.get(h, ""))
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = min(48, max(12, len(str(headers[c - 1])) + 4))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_text_sheet_bytes(title: str, body: str, sheet_name: str = "Export") -> bytes:
    """Create a single-sheet Excel file with a title row and body text."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.cell(row=1, column=1, value=_normalize_cell(title))
    ws.cell(row=2, column=1, value=_normalize_cell(body))
    ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    hdr = PatternFill(start_color=SF_HEADER_FILL, end_color=SF_HEADER_FILL, fill_type="solid")
    ws.cell(row=1, column=1).fill = hdr
    for cell in (ws["A1"], ws["A2"]):
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sanitize_filename(name: str) -> str:
    """Return a filesystem-safe lowercase stem."""
    return re.sub(r"[^\w\-]", "_", name).lower()


def export_to_markdown(content: str, agent_name: str) -> bytes:
    """Prefix agent output with a title block for download."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    header = (
        f"# Salesforce QA Agent — {agent_name.title()}\n"
        f"> Generated: {timestamp}\n\n---\n\n"
    )
    return (header + content).encode("utf-8")


def split_markdown_sections(content: str) -> list[dict[str, str]]:
    """Split markdown into sections keyed by heading.

    A new section begins on every ATX heading line (``#`` to ``######``) that is
    *not* inside a fenced code block. Anything before the first heading is kept
    as a ``(Preamble)`` section so no content is lost. Each returned dict has
    ``Section Title`` and ``Markdown Content`` keys; the heading line itself is
    preserved at the top of the content for round-tripping.
    """
    sections: list[dict[str, str]] = []
    current_title = "(Preamble)"
    current_lines: list[str] = []
    in_fence = False

    def _flush() -> None:
        text = "\n".join(current_lines).strip("\n")
        if text or sections:
            sections.append(
                {"Section Title": current_title, "Markdown Content": text}
            )

    for raw_line in (content or "").splitlines():
        if _FENCE_RE.match(raw_line):
            in_fence = not in_fence
            current_lines.append(raw_line)
            continue
        m = _HEADING_RE.match(raw_line) if not in_fence else None
        if m:
            _flush()
            current_title = m.group(2).strip() or "(Untitled)"
            current_lines = [raw_line]
        else:
            current_lines.append(raw_line)
    _flush()

    cleaned: list[dict[str, str]] = []
    for sec in sections:
        if sec["Section Title"] == "(Preamble)" and not sec["Markdown Content"].strip():
            continue
        cleaned.append(sec)
    if not cleaned and (content or "").strip():
        cleaned.append({"Section Title": "(Content)", "Markdown Content": content})
    return cleaned


_SEP_RE = re.compile(r"^[\s|:\-]+$")


def parse_all_markdown_tables(content: str) -> list[dict]:
    """Extract every GFM pipe table from *content*, tagged with its nearest ATX heading.

    Returns a list of dicts::

        [
            {
                "heading": "Test Cases",          # nearest ## heading before the table
                "headers": ["TC ID", "Summary"],  # column names from the header row
                "rows": [["TC_001", "…"], …],     # list of cell lists (strings)
            },
            …
        ]

    Tables inside fenced code blocks are skipped. Escaped pipes (``\\|``) inside
    cells are temporarily substituted so splitting on ``|`` still works.
    """
    _PIPE_ESC = "\x00PIPE\x00"
    tables: list[dict] = []
    current_heading = "(Untitled)"
    in_fence = False
    lines = (content or "").splitlines()
    i = 0

    def _parse_row(raw: str) -> list[str]:
        escaped = raw.replace("\\|", _PIPE_ESC)
        cells = [c.replace(_PIPE_ESC, "|").strip() for c in escaped.strip("|").split("|")]
        return cells

    def _is_separator(raw: str) -> bool:
        stripped = raw.strip()
        if not stripped or "|" not in stripped:
            return False
        inner = stripped.strip("|")
        return bool(_SEP_RE.match(inner))

    def _is_pipe_row(raw: str) -> bool:
        stripped = raw.strip()
        return "|" in stripped and stripped.count("|") >= 1

    while i < len(lines):
        raw = lines[i]

        # Track fenced code blocks
        if _FENCE_RE.match(raw):
            in_fence = not in_fence
            i += 1
            continue

        if in_fence:
            i += 1
            continue

        # Track headings
        m = _HEADING_RE.match(raw)
        if m:
            current_heading = m.group(2).strip() or "(Untitled)"
            i += 1
            continue

        # Look for the start of a pipe table: pipe row followed by separator row
        if _is_pipe_row(raw) and i + 1 < len(lines) and _is_separator(lines[i + 1]):
            headers = _parse_row(raw)
            i += 2  # skip header row + separator row
            data_rows: list[list[str]] = []
            while i < len(lines):
                data_line = lines[i]
                if not _is_pipe_row(data_line) or _is_separator(data_line):
                    break
                data_rows.append(_parse_row(data_line))
                i += 1
            if headers:
                tables.append({
                    "heading": current_heading,
                    "headers": headers,
                    "rows": data_rows,
                })
            continue

        i += 1

    return tables


def export_to_csv(content: str) -> bytes:
    """Return CSV bytes built from every GFM pipe table in *content*.

    - **One table**: flat CSV — header row = table column names, one data row per record.
    - **Multiple tables**: all tables in one file, each preceded by a ``# Section:``
      comment row and separated by a blank row.
    - **No tables** (prose-only output): falls back to section-dump (``Section Title``,
      ``Markdown Content``) so the file is never empty.
    """
    tables = parse_all_markdown_tables(content)
    buf = io.StringIO()
    writer = csv.writer(buf)

    if not tables:
        # Fallback: section-based dump
        rows = split_markdown_sections(content)
        if not rows:
            rows = [{"Section Title": "(Empty)", "Markdown Content": ""}]
        writer.writerow(["Section Title", "Markdown Content"])
        for row in rows:
            writer.writerow([
                _normalize_cell(row.get("Section Title", "")),
                _normalize_cell(row.get("Markdown Content", "")),
            ])
        return buf.getvalue().encode("utf-8")

    for idx, tbl in enumerate(tables):
        if idx > 0:
            writer.writerow([])  # blank separator row
        if len(tables) > 1:
            writer.writerow([f"# Section: {tbl['heading']}"])
        writer.writerow(tbl["headers"])
        for row in tbl["rows"]:
            ncols = len(tbl["headers"])
            padded = (row + [""] * ncols)[:ncols]
            writer.writerow([_normalize_cell(v) for v in padded])

    return buf.getvalue().encode("utf-8")


def export_to_excel(content: str, agent_name: str) -> bytes:
    """Build an .xlsx from the GFM pipe tables in *content*.

    - **Tables found**: one worksheet per table, named after the section heading.
      Row 1 = real column headers (Salesforce-blue); subsequent rows = data rows.
    - **No tables**: falls back to section-dump on a single sheet.
    """
    tables = parse_all_markdown_tables(content)

    if not tables:
        rows = split_markdown_sections(content)
        sheet = (agent_name or "export")[:31] or "Export"
        if not rows:
            rows = [{"Section Title": "(Empty)", "Markdown Content": ""}]
        return export_workbook_bytes({sheet: rows})

    wb = Workbook()
    first = True
    seen_names: dict[str, int] = {}

    for tbl in tables:
        # Build a unique, valid sheet name (max 31 chars)
        raw_name = (tbl["heading"] or "Sheet")[:28].strip()
        if not raw_name:
            raw_name = "Sheet"
        count = seen_names.get(raw_name, 0)
        seen_names[raw_name] = count + 1
        sheet_name = raw_name if count == 0 else f"{raw_name[:25]} ({count})"

        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        headers = tbl["headers"]
        ncols = len(headers)

        # Write + style header row
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, ncols)

        # Write data rows
        for r, row_cells in enumerate(tbl["rows"], start=2):
            padded = (row_cells + [""] * ncols)[:ncols]
            for c, val in enumerate(padded, start=1):
                cell = ws.cell(row=r, column=c, value=_normalize_cell(val))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-size columns (capped at 60)
        for c, h in enumerate(headers, start=1):
            col_vals = [h] + [(tbl["rows"][r][c - 1] if c - 1 < len(tbl["rows"][r]) else "") for r in range(len(tbl["rows"]))]
            max_len = max((len(str(v)) for v in col_vals), default=8)
            ws.column_dimensions[get_column_letter(c)].width = min(60, max(12, max_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_PDF_CSS = f"""
@page {{ size: A4; margin: 18mm 16mm 20mm 16mm; @frame footer {{
    -pdf-frame-content: footer; bottom: 8mm; margin-left: 16mm; margin-right: 16mm; height: 10mm;
}} }}
body {{ font-family: Helvetica, Arial, sans-serif; font-size: 10.5pt; color: #1a1a1a; line-height: 1.45; }}
h1 {{ color: {SF_HEADER_HEX}; font-size: 22pt; border-bottom: 2pt solid {SF_HEADER_HEX};
       padding-bottom: 4pt; margin-top: 0; }}
h2 {{ color: {SF_HEADER_HEX}; font-size: 16pt; margin-top: 18pt; border-bottom: 0.5pt solid #cfd9e3; padding-bottom: 2pt; }}
h3 {{ color: #16325c; font-size: 13pt; margin-top: 14pt; }}
h4, h5, h6 {{ color: #16325c; margin-top: 10pt; }}
p {{ margin: 6pt 0; }}
ul, ol {{ margin: 6pt 0 6pt 18pt; }}
li {{ margin-bottom: 2pt; }}
code {{ font-family: Courier, monospace; background: #f4f6f9; padding: 1pt 3pt; border-radius: 2pt; }}
pre {{ font-family: Courier, monospace; background: #f4f6f9; padding: 8pt; border: 0.5pt solid #d8dde6;
       border-radius: 3pt; white-space: pre-wrap; font-size: 9pt; }}
blockquote {{ border-left: 3pt solid {SF_HEADER_HEX}; padding-left: 8pt; color: #4a5568; margin: 8pt 0; }}
table {{ border-collapse: collapse; width: 100%; margin: 8pt 0; }}
th {{ background: {SF_HEADER_HEX}; color: #ffffff; padding: 5pt 6pt; text-align: left;
      border: 0.5pt solid {SF_HEADER_HEX}; font-size: 9.5pt; }}
td {{ padding: 4pt 6pt; border: 0.5pt solid #d8dde6; vertical-align: top; font-size: 9.5pt; }}
.header-banner {{ color: #6b7280; font-size: 9pt; margin-bottom: 12pt; }}
.footer {{ color: #94a3b8; font-size: 8pt; text-align: center; }}
hr {{ border: 0; border-top: 0.5pt solid #cfd9e3; margin: 12pt 0; }}
"""


def export_to_pdf(content: str, agent_name: str) -> bytes:
    """Render markdown content to a styled PDF document.

    Markdown is converted to HTML (with GitHub-flavoured table support) and
    then rendered with xhtml2pdf/pisa. The output uses Salesforce-blue
    accents to match the rest of the app.
    """
    import markdown as md
    from xhtml2pdf import pisa

    # Normalize all <br>-style sentinels (raw or HTML-entity escaped) to a
    # single canonical <br/> tag so xhtml2pdf renders them as real line breaks
    # inside table cells.
    normalized_content = _BR_SENTINEL_RE.sub("<br/>", content or "")

    html_body = md.markdown(
        normalized_content,
        extensions=["tables", "fenced_code", "sane_lists", "toc"],
        output_format="html5",
    )
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    title = f"Salesforce QA Agent — {(agent_name or 'export').replace('_', ' ').title()}"
    safe_title = _html.escape(title)
    safe_agent = _html.escape(agent_name or "export")
    safe_ts = _html.escape(timestamp)

    document = f"""<!DOCTYPE html>
<html><head><meta charset=\"utf-8\"><style>{_PDF_CSS}</style></head>
<body>
<div id=\"footer\" class=\"footer\">QA Studio • {safe_agent} • {safe_ts} • Page <pdf:pagenumber/> of <pdf:pagecount/></div>
<h1>{safe_title}</h1>
<div class=\"header-banner\">Generated {safe_ts}</div>
<hr/>
{html_body}
</body></html>"""

    buf = io.BytesIO()
    result = pisa.CreatePDF(src=io.StringIO(document), dest=buf, encoding="utf-8")
    if result.err:
        raise RuntimeError(f"PDF generation failed ({result.err} errors)")
    return buf.getvalue()
