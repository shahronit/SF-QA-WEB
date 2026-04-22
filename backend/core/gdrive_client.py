"""Google Drive OAuth + file reader.

Provides:
  - OAuth 2.0 helpers (`build_auth_url`, `exchange_code`, `refresh_access_token`)
  - `GDriveClient` class that downloads + extracts text from Google Docs,
    Sheets, PDFs, Excel files, CSV files and plain text — keyed off a
    per-user credentials dict produced by the OAuth dance.

Designed to mirror the session pattern used by `jira_client.JiraClient` so
the FastAPI router layer can treat both integrations uniformly.
"""

from __future__ import annotations

import csv
import io
import logging
import re
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from typing import Any

from config import settings

log = logging.getLogger(__name__)

# Read-only scope is enough for fetching metadata + content.
SCOPES = ["https://www.googleapis.com/auth/drive.readonly", "openid", "email"]

# Authorization endpoints
_AUTH_URI = "https://accounts.google.com/o/oauth2/v2/auth"
_TOKEN_URI = "https://oauth2.googleapis.com/token"
_USERINFO_URI = "https://openidconnect.googleapis.com/v1/userinfo"


# ---------------------------------------------------------------------------
# OAuth helpers
# ---------------------------------------------------------------------------

def _client_config_dict() -> dict[str, Any]:
    """Build the InstalledAppFlow-style client config dict from app settings."""
    return {
        "web": {
            "client_id": settings.GOOGLE_OAUTH_CLIENT_ID,
            "client_secret": settings.GOOGLE_OAUTH_CLIENT_SECRET,
            "auth_uri": _AUTH_URI,
            "token_uri": _TOKEN_URI,
            "redirect_uris": [settings.GOOGLE_OAUTH_REDIRECT_URI],
        }
    }


def is_oauth_configured() -> bool:
    """Return True when both client id + secret are present in env."""
    return bool(settings.GOOGLE_OAUTH_CLIENT_ID and settings.GOOGLE_OAUTH_CLIENT_SECRET)


def build_auth_url(state: str) -> str:
    """Return the Google consent URL the user should be redirected to.

    Uses ``access_type=offline`` + ``prompt=consent`` so the first hit always
    returns a refresh token — necessary because Google only emits a refresh
    token on the first user consent unless prompted again.
    """
    if not is_oauth_configured():
        raise RuntimeError(
            "Google OAuth is not configured. Set GOOGLE_OAUTH_CLIENT_ID and "
            "GOOGLE_OAUTH_CLIENT_SECRET environment variables."
        )
    from google_auth_oauthlib.flow import Flow  # lazy import so missing dep is friendlier

    flow = Flow.from_client_config(
        _client_config_dict(),
        scopes=SCOPES,
        redirect_uri=settings.GOOGLE_OAUTH_REDIRECT_URI,
    )
    auth_url, _ = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
        state=state,
    )
    return auth_url


def exchange_code(code: str) -> dict[str, Any]:
    """Swap an authorization code for tokens + user email.

    Returns a session-shaped dict with all fields needed to re-build a
    GDriveClient and to refresh the access token later.
    """
    if not is_oauth_configured():
        raise RuntimeError("Google OAuth is not configured.")

    from google_auth_oauthlib.flow import Flow

    flow = Flow.from_client_config(
        _client_config_dict(),
        scopes=SCOPES,
        redirect_uri=settings.GOOGLE_OAUTH_REDIRECT_URI,
    )
    flow.fetch_token(code=code)
    creds = flow.credentials

    email = _fetch_userinfo_email(creds.token)

    expires_at = (
        creds.expiry.replace(tzinfo=timezone.utc).isoformat()
        if creds.expiry
        else (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
    )

    return {
        "access_token": creds.token,
        "refresh_token": creds.refresh_token,
        "token_uri": creds.token_uri or _TOKEN_URI,
        "client_id": creds.client_id or settings.GOOGLE_OAUTH_CLIENT_ID,
        "client_secret": creds.client_secret or settings.GOOGLE_OAUTH_CLIENT_SECRET,
        "scopes": list(creds.scopes or SCOPES),
        "expires_at": expires_at,
        "email": email,
        "connected_at": datetime.now(timezone.utc).isoformat(),
    }


def _fetch_userinfo_email(access_token: str) -> str:
    """Return the user's email from the OIDC userinfo endpoint."""
    import urllib.request

    req = urllib.request.Request(_USERINFO_URI)
    req.add_header("Authorization", f"Bearer {access_token}")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            import json
            data = json.loads(resp.read().decode())
            return data.get("email", "")
    except Exception:  # noqa: BLE001
        return ""


def _credentials_from_session(session: dict[str, Any]):
    """Build a `google.oauth2.credentials.Credentials` from a stored session."""
    from google.oauth2.credentials import Credentials

    return Credentials(
        token=session.get("access_token"),
        refresh_token=session.get("refresh_token"),
        token_uri=session.get("token_uri", _TOKEN_URI),
        client_id=session.get("client_id", settings.GOOGLE_OAUTH_CLIENT_ID),
        client_secret=session.get("client_secret", settings.GOOGLE_OAUTH_CLIENT_SECRET),
        scopes=session.get("scopes", SCOPES),
    )


def refresh_if_needed(session: dict[str, Any]) -> dict[str, Any]:
    """Refresh the access token if expired and return the updated session.

    Returns the same dict (possibly updated in-place + with a new
    ``access_token`` / ``expires_at``) so the caller can re-persist it.
    """
    expires_at_raw = session.get("expires_at")
    needs_refresh = True
    if expires_at_raw:
        try:
            exp = datetime.fromisoformat(expires_at_raw.replace("Z", "+00:00"))
            # Refresh 60 seconds before actual expiry to absorb clock skew.
            needs_refresh = exp <= datetime.now(timezone.utc) + timedelta(seconds=60)
        except ValueError:
            needs_refresh = True

    if not needs_refresh:
        return session

    from google.auth.transport.requests import Request

    creds = _credentials_from_session(session)
    if not creds.refresh_token:
        # Cannot refresh without a refresh token; let the caller handle re-auth.
        return session
    creds.refresh(Request())
    session["access_token"] = creds.token
    if creds.expiry:
        session["expires_at"] = creds.expiry.replace(tzinfo=timezone.utc).isoformat()
    return session


# ---------------------------------------------------------------------------
# URL / file-id parsing
# ---------------------------------------------------------------------------

# Drive file ids are URL-safe base64-ish strings, generally 25-44 chars.
_FILE_ID_RE = re.compile(r"[A-Za-z0-9_-]{20,}")


def extract_file_id(url_or_id: str) -> str | None:
    """Extract a Drive file id from a URL or return the input if it is one.

    Accepts:
      - raw file ids
      - https://docs.google.com/document/d/<id>/edit
      - https://docs.google.com/spreadsheets/d/<id>/...
      - https://docs.google.com/presentation/d/<id>/...
      - https://drive.google.com/file/d/<id>/view
      - https://drive.google.com/open?id=<id>
      - https://drive.google.com/uc?id=<id>
    """
    if not url_or_id:
        return None
    text = url_or_id.strip()

    # Looks like a bare id already.
    if "://" not in text and "/" not in text and _FILE_ID_RE.fullmatch(text):
        return text

    # Try the path-component format: .../d/<id>/...
    path_match = re.search(r"/d/([A-Za-z0-9_-]{20,})", text)
    if path_match:
        return path_match.group(1)

    # Try the query-string format: ?id=<id>
    try:
        parsed = urllib.parse.urlsplit(text)
        qs = urllib.parse.parse_qs(parsed.query)
        if "id" in qs and qs["id"]:
            candidate = qs["id"][0]
            if _FILE_ID_RE.fullmatch(candidate):
                return candidate
    except ValueError:
        pass

    return None


# ---------------------------------------------------------------------------
# GDriveClient
# ---------------------------------------------------------------------------

# Mime types we know how to handle natively.
_GDOC_MIME = "application/vnd.google-apps.document"
_GSHEET_MIME = "application/vnd.google-apps.spreadsheet"
_GSLIDES_MIME = "application/vnd.google-apps.presentation"
_PDF_MIME = "application/pdf"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLS_MIME = "application/vnd.ms-excel"
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

# Cap how much text we hand back per file so a giant PDF cannot blow up the
# JSON response. Adjust upward if you need full-content RAG ingestion.
MAX_TEXT_CHARS = 200_000


class GDriveClient:
    """Per-user wrapper around the Drive v3 API service."""

    def __init__(self, session: dict[str, Any]) -> None:
        from googleapiclient.discovery import build  # lazy import

        self.session = refresh_if_needed(session)
        creds = _credentials_from_session(self.session)
        self.service = build("drive", "v3", credentials=creds, cache_discovery=False)

    # ------------------------------------------------------------------
    # Metadata
    # ------------------------------------------------------------------

    def get_metadata(self, file_id: str) -> dict[str, Any]:
        """Return basic metadata about a file."""
        meta = (
            self.service.files()
            .get(
                fileId=file_id,
                fields="id, name, mimeType, size, modifiedTime, webViewLink, owners",
                supportsAllDrives=True,
            )
            .execute()
        )
        return {
            "id": meta.get("id"),
            "name": meta.get("name"),
            "mime_type": meta.get("mimeType"),
            "size": int(meta["size"]) if meta.get("size") else None,
            "modified_time": meta.get("modifiedTime"),
            "web_view_link": meta.get("webViewLink"),
            "owners": [o.get("emailAddress") for o in (meta.get("owners") or [])],
        }

    # ------------------------------------------------------------------
    # Content readers (one per mime category)
    # ------------------------------------------------------------------

    def read_file(self, url_or_id: str) -> dict[str, Any]:
        """Read a file's text content, dispatching by mime type.

        Always returns a dict — never raises — so callers can collect
        partial successes.
        """
        file_id = extract_file_id(url_or_id)
        if not file_id:
            return {
                "input": url_or_id,
                "id": None,
                "name": None,
                "text": "",
                "error": "INVALID_URL",
                "message": f"Could not extract a Drive file id from: {url_or_id!r}",
            }

        try:
            meta = self.get_metadata(file_id)
        except Exception as exc:  # noqa: BLE001
            return self._error_result(file_id, url_or_id, exc, "METADATA_FAILED")

        mime = meta.get("mime_type", "")

        try:
            if mime == _GDOC_MIME:
                text = self._export_native(file_id, "text/plain")
            elif mime == _GSHEET_MIME:
                text = self._export_native(file_id, "text/csv")
            elif mime == _GSLIDES_MIME:
                text = self._export_native(file_id, "text/plain")
            elif mime == _PDF_MIME:
                text = self._read_pdf(file_id)
            elif mime == _XLSX_MIME:
                text = self._read_xlsx(file_id)
            elif mime in {"text/csv", _XLS_MIME}:
                text = self._read_csv(file_id)
            elif mime == _DOCX_MIME:
                text = self._read_docx(file_id)
            elif mime.startswith("text/"):
                raw = self._download_bytes(file_id)
                text = raw.decode("utf-8", errors="replace")
            else:
                return {
                    **meta,
                    "input": url_or_id,
                    "text": "",
                    "error": "UNSUPPORTED_MIME_TYPE",
                    "message": f"Unsupported mime type: {mime}",
                }
        except Exception as exc:  # noqa: BLE001
            return self._error_result(file_id, url_or_id, exc, "READ_FAILED", meta=meta)

        # Truncate to keep the response payload bounded.
        truncated = False
        if text and len(text) > MAX_TEXT_CHARS:
            text = text[:MAX_TEXT_CHARS]
            truncated = True

        return {
            **meta,
            "input": url_or_id,
            "text": text,
            "truncated": truncated,
        }

    def read_many(
        self,
        urls_or_ids: list[str],
        max_workers: int = 4,
    ) -> list[dict[str, Any]]:
        """Read multiple files in parallel; preserves input order."""
        # Deduplicate by extracted file id (or original input if none) so we
        # don't double-fetch when the same Drive URL appears in many comments.
        seen: dict[str, str] = {}
        ordered_inputs: list[str] = []
        for raw in urls_or_ids:
            key = extract_file_id(raw) or raw
            if key not in seen:
                seen[key] = raw
                ordered_inputs.append(raw)

        results: dict[str, dict[str, Any]] = {}
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(self.read_file, url): url
                for url in ordered_inputs
            }
            for future in as_completed(futures):
                url = futures[future]
                try:
                    results[url] = future.result()
                except Exception as exc:  # noqa: BLE001
                    results[url] = {
                        "input": url,
                        "text": "",
                        "error": "READ_FAILED",
                        "message": str(exc),
                    }

        return [results[url] for url in ordered_inputs]

    # ------------------------------------------------------------------
    # Lower-level download primitives
    # ------------------------------------------------------------------

    def _download_bytes(self, file_id: str) -> bytes:
        """Download a binary file via files().get_media."""
        from googleapiclient.http import MediaIoBaseDownload

        request = self.service.files().get_media(fileId=file_id, supportsAllDrives=True)
        buffer = io.BytesIO()
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buffer.getvalue()

    def _export_native(self, file_id: str, mime_type: str) -> str:
        """Export a Google-native doc (Docs/Sheets/Slides) as text."""
        request = self.service.files().export_media(
            fileId=file_id, mimeType=mime_type
        )
        # export_media returns the bytes directly.
        data = request.execute()
        if isinstance(data, bytes):
            return data.decode("utf-8", errors="replace")
        return str(data)

    # ------------------------------------------------------------------
    # Format-specific extractors
    # ------------------------------------------------------------------

    def _read_pdf(self, file_id: str) -> str:
        """Extract text from a PDF; tries pypdf first, falls back to pdfplumber."""
        raw = self._download_bytes(file_id)

        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(raw))
            chunks: list[str] = []
            for page in reader.pages:
                try:
                    chunks.append(page.extract_text() or "")
                except Exception:  # noqa: BLE001
                    continue
            text = "\n\n".join(chunks).strip()
            if text:
                return text
        except Exception as exc:  # noqa: BLE001
            log.warning("pypdf failed for %s, trying pdfplumber: %s", file_id, exc)

        # Fallback for PDFs pypdf chokes on (scanned, complex layout).
        try:
            import pdfplumber  # type: ignore

            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                pages = [p.extract_text() or "" for p in pdf.pages]
            return "\n\n".join(pages).strip()
        except Exception as exc:  # noqa: BLE001
            log.warning("pdfplumber also failed for %s: %s", file_id, exc)
            return ""

    def _read_xlsx(self, file_id: str) -> str:
        """Read an Excel file and return all sheets as TSV-style text."""
        from openpyxl import load_workbook

        raw = self._download_bytes(file_id)
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        out: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            out.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                out.append("\t".join(cells))
            out.append("")
        return "\n".join(out)

    def _read_csv(self, file_id: str) -> str:
        """Read a CSV (or csv-as-xls) file as text."""
        raw = self._download_bytes(file_id)
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("utf-8", errors="replace")
        # Round-trip through csv to normalise encodings/line endings.
        try:
            reader = csv.reader(io.StringIO(text))
            return "\n".join(",".join(row) for row in reader)
        except Exception:  # noqa: BLE001
            return text

    def _read_docx(self, file_id: str) -> str:
        """Best-effort read of a .docx via python-docx (already a transitive dep)."""
        raw = self._download_bytes(file_id)
        try:
            from docx import Document  # type: ignore

            doc = Document(io.BytesIO(raw))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as exc:  # noqa: BLE001
            log.warning("docx parse failed for %s: %s", file_id, exc)
            return ""

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _error_result(
        file_id: str,
        url: str,
        exc: Exception,
        code: str,
        meta: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Format a failure result with as much context as we have."""
        base = meta.copy() if meta else {"id": file_id, "name": None, "mime_type": None}
        base.update({
            "input": url,
            "text": "",
            "error": code,
            "message": f"{type(exc).__name__}: {exc}",
        })
        return base
